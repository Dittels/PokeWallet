from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from openpyxl import load_workbook, Workbook
import io
import datetime
import os

from scraper import URLScrape, CardUpdate




print('Collection file:')
print('1: use existing')
print('2: create new')
menu0 = int(input())
x = True

if(menu0 == 1):
    print('enter file name')
    colName = str(input())
    #load excel file
    workbook = load_workbook(filename="CollectionSheet/"+colName+".xlsx") 
    #open workbook
    sheet = workbook.active
    sheet["A1"] = "Card Name"
    sheet["B1"] = "Card Market Price"
    sheet["C1"] = "Last Updated"
    sheet["D1"] = "Card URL"
    sheet["F1"] = "Total"

elif(menu0 == 2):
    print('enter desired file name')
    colName = str(input())
    print(colName)
    workbook = Workbook()
    workbook.save("CollectionSheet/"+colName+".xlsx")
    #open workbook
    sheet = workbook.active
    sheet["A1"] = "Card Name"
    sheet["B1"] = "Card Market Price"
    sheet["C1"] = "Last Updated"
    sheet["D1"] = "Card URL"
    sheet["F1"] = "Total"

else:
    print('invalid input')
    x= False


while x:
    print('Select Action (only enter number):')
    print('1: Add new card')
    print('2: update current prices')
    print('3: Add bulk cards')
    print('4: exit')
    menu1 = int(input())
    if(menu1 == 1):
        print("Enter Card URL:")
        cardUrl = str(input())
        cardName, CardPrice = URLScrape(cardUrl)
        sheet.append([cardName, CardPrice,  datetime.datetime.now().strftime("%m-%d-%Y"), cardUrl])
        workbook.save("CollectionSheet/"+colName+".xlsx")

        
    elif(menu1 == 2):
        print("we're getting there")
    elif(menu1 == 3):
        with open("urls.txt", 'r') as obj:
            urls = obj.readlines()

        for cardUrl in urls:
            cardName, CardPrice = URLScrape(cardUrl.replace("\n", ""))
            sheet.append([cardName, CardPrice,  datetime.datetime.now().strftime("%m-%d-%Y"), cardUrl])
        workbook.save("CollectionSheet/"+colName+".xlsx")
    elif(menu1 == 4):
        print('bye bye')
        x = False
    else:
        print('invalid input')

sheet['F2']="=SUM(B:B)"
workbook.save("CollectionSheet/"+colName+".xlsx")

workbook.close()